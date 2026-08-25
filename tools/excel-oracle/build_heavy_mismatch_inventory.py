from __future__ import annotations

import datetime
import hashlib
import json
import os
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import to_excel

import formualizer as fz

WORKBOOK_PATH = r"C:\Users\OXK0A0A\Downloads\Fossil_EstimatingTemplate_2026-08_21_A.xlsx"
BASELINE_PATH = r"C:\rust_engines\formualizer-v0.8.4\docs\issue-solutions\data\latest-upstream-heavy-baseline.json"
EXCEL_SNAPSHOT = os.environ["EXCEL_SNAPSHOT"]
OUTPUT_PATH = Path(
    r"C:\rust_engines\formualizer-v0.8.4\docs\issue-solutions\data\heavy-formualizer-excel-mismatch-inventory.json"
)

ERROR_KIND = {
    "#NULL!": "Null",
    "#DIV/0!": "Div",
    "#VALUE!": "Value",
    "#REF!": "Ref",
    "#NAME?": "Name",
    "#NUM!": "Num",
    "#N/A": "Na",
    "#SPILL!": "Spill",
    "#CALC!": "Calc",
}


def normalize_address(address: str) -> str:
    if "!" not in address:
        return address.replace("$", "").upper()
    sheet, cell = address.split("!", 1)
    return f"{sheet.strip(chr(39)).replace(chr(39) * 2, chr(39)).upper()}!{cell.replace('$', '').upper()}"


def excel_value(value: Any) -> dict[str, Any]:
    if value is None:
        return {"kind": "empty", "value": None}
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return {"kind": "number", "value": to_excel(value)}
    if isinstance(value, str):
        upper = value.upper()
        if upper in ERROR_KIND:
            return {"kind": "error", "error_kind": ERROR_KIND[upper], "display": upper}
        return {"kind": "string", "value": value}
    if isinstance(value, bool):
        return {"kind": "boolean", "value": value}
    if isinstance(value, (int, float)):
        return {"kind": "number", "value": value}
    return {"kind": type(value).__name__, "value": str(value)}


def fz_value(value: Any) -> dict[str, Any]:
    if value is None:
        return {"kind": "empty", "value": None}
    if isinstance(value, dict) and value.get("type") == "Error":
        return {"kind": "error", "error_kind": value.get("kind"), "value": value}
    if isinstance(value, bool):
        return {"kind": "boolean", "value": value}
    if isinstance(value, (int, float)):
        return {"kind": "number", "value": value}
    if isinstance(value, str):
        return {"kind": "string", "value": value}
    if isinstance(value, list):
        return {"kind": "array", "value": value, "shape": [len(value), max((len(r) for r in value), default=0)]}
    return {"kind": type(value).__name__, "value": str(value)}


def classify(excel: dict[str, Any], engine: dict[str, Any]) -> str:
    if excel["kind"] == "error" and engine["kind"] == "error":
        return "match" if excel.get("error_kind") == engine.get("error_kind") else "error_kind_mismatch"
    if excel["kind"] == "number" and engine["kind"] == "number":
        delta = abs(float(excel["value"]) - float(engine["value"]))
        return "match" if delta <= 1e-9 else "numeric_value_mismatch"
    if excel["kind"] == engine["kind"] and excel.get("value") == engine.get("value"):
        return "match"
    if excel["kind"] == "number" and engine["kind"] == "error":
        return f"excel_numeric_formualizer_{str(engine.get('error_kind', 'error')).lower()}_error"
    if excel["kind"] == "error" and engine["kind"] == "number":
        return f"excel_error_formualizer_numeric_{str(excel.get('error_kind', 'error')).lower()}"
    return "other_type_or_value_mismatch"


def formula_features(formula: str) -> list[str]:
    upper = formula.upper()
    features = []
    for name, label in [
        ("INDIRECT", "dynamic_reference"),
        ("OFFSET", "dynamic_reference"),
        ("INDEX", "index_reference"),
        ("IF", "conditional"),
        ("IFS", "conditional"),
        ("CHOOSE", "conditional"),
        ("SWITCH", "conditional"),
        ("FILTER", "array_spill"),
        ("SEQUENCE", "array_spill"),
        ("VSTACK", "array_spill"),
        ("UNIQUE", "array_spill"),
        ("MAP", "unsupported_or_xlfn"),
        ("LAMBDA", "unsupported_or_xlfn"),
    ]:
        if re.search(rf"(?<![A-Z0-9_]){name}(?![A-Z0-9_])", upper):
            if label not in features:
                features.append(label)
    if ":" in formula:
        features.append("range_reference")
    return features


def main() -> None:
    baseline = json.loads(Path(BASELINE_PATH).read_text(encoding="utf-8"))
    static_members = {
        normalize_address(address)
        for address in baseline["steps"][0]["main_passes"][0]["changed_member_addresses"]
    }
    runtime_samples = {
        normalize_address(address)
        for address in baseline["steps"][0]["main_runtime"]["volatile_member_samples"]
    }
    runtime_samples.update(
        normalize_address(address)
        for address in baseline["steps"][0]["main_runtime"]["dynamic_member_samples"]
    )

    cfg = fz.EvaluationConfig()
    cfg.enable_parallel = True
    cfg.cycle_detection = "runtime"
    cfg.cycle_policy = "iterate"
    cfg.iterate_max_iterations = 100
    cfg.iterate_max_change = 0.001
    workbook = fz.Workbook.load_path(WORKBOOK_PATH, backend="calamine", config=fz.WorkbookConfig(eval_config=cfg))
    workbook.set_value("Inputs", 7, 6, 300)
    workbook.evaluate_all()
    engine_snapshot = dict(workbook.formula_output_snapshot())

    excel_values = openpyxl.load_workbook(EXCEL_SNAPSHOT, data_only=True, read_only=True)
    excel_formulas = openpyxl.load_workbook(EXCEL_SNAPSHOT, data_only=False, read_only=True)
    engine_by_address = {
        normalize_address(address): (address, engine_raw)
        for address, engine_raw in engine_snapshot.items()
        if "!" in address
    }
    mismatches = []
    counts: dict[str, int] = {}
    main_counts: dict[str, int] = {}
    main_mismatches = []
    for formula_sheet in excel_formulas.worksheets:
        value_sheet = excel_values[formula_sheet.title]
        for row_index, (formula_row, value_row) in enumerate(
            zip(formula_sheet.iter_rows(), value_sheet.iter_rows()), 1
        ):
            for column_index, (formula_cell, value_cell) in enumerate(
                zip(formula_row, value_row), 1
            ):
                normalized = normalize_address(
                    f"{formula_sheet.title}!{get_column_letter(column_index)}{row_index}"
                )
                if normalized not in engine_by_address:
                    continue
                address, engine_raw = engine_by_address[normalized]
                formula_raw = formula_cell.value
                formula = getattr(formula_raw, "text", formula_raw)
                if not isinstance(formula, str):
                    formula = str(formula)
                excel = excel_value(value_cell.value)
                engine = fz_value(engine_raw)
                category = classify(excel, engine)
                if category == "match":
                    continue
                item = {
                    "address": address,
                    "formula": formula,
                    "excel": excel,
                    "formualizer": engine,
                    "category": category,
                    "inside_static_scc": normalized in static_members,
                    "runtime_live_membership": "known_in_runtime_sample_only" if normalized in runtime_samples else "unknown",
                    "runtime_live_membership_proven": False,
                    "formula_features": formula_features(formula),
                    "internal_outgoing_edge_families": "not available in current public diagnostic artifact",
                    "internal_incoming_edge_families": "not available in current public diagnostic artifact",
                }
                mismatches.append(item)
                counts[category] = counts.get(category, 0) + 1
                if item["inside_static_scc"]:
                    main_mismatches.append(item)
                    main_counts[category] = main_counts.get(category, 0) + 1

    excel_values.close()
    excel_formulas.close()
    OUTPUT_PATH.write_text(
        json.dumps(
            {
                "schema": "formualizer.heavy-excel-mismatch-inventory/v1",
                "workbook": Path(WORKBOOK_PATH).name,
                "excel_snapshot": EXCEL_SNAPSHOT,
                "input": "Inputs!F7=300",
                "formualizer_static_scc_member_count": len(static_members),
                "formualizer_runtime_live_member_count": baseline["steps"][0]["main_runtime"]["live_cycle_member_count"],
                "runtime_live_full_address_set_available": False,
                "all_formula_mismatch_counts": counts,
                "main_scc_mismatch_counts": main_counts,
                "all_formula_mismatch_count": len(mismatches),
                "main_scc_mismatch_count": len(main_mismatches),
                "mismatches": mismatches,
                "main_scc_mismatches": main_mismatches,
                "notes": [
                    "Excel values come from a fresh COM-recalculated copy with Inputs!F7=300.",
                    "Formualizer values come from the current diagnostic evaluator with Inputs!F7=300.",
                    "Runtime-live membership is unknown except for explicit prior sample addresses.",
                    "Formula feature labels are lexical diagnostics, not edge-origin proof.",
                ],
            },
            indent=2,
            default=str,
        )
        + "\n",
        encoding="utf-8",
    )
    print(f"Generated {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
